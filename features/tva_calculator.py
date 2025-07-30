import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Logic for calculating HT and TVA
def calculate_ht_tva(ttc, tva_rate):
    ht = round(ttc / (1 + tva_rate / 100), 2)
    tva = round(ttc - ht, 2)
    return ht, tva

# Function to export DataFrame to Excel in memory
def export_to_excel(df, enterprise_name, date_str):
    wb = Workbook()
    ws = wb.active
    ws.title = "TVA Report"

    for i in range(1, 100):
        ws.row_dimensions[i].height = 22
    for col in range(2, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    grey_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def write_section(start_row, title, data, is_client, first_col_header):
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
        ws.cell(row=start_row, column=2).value = title
        ws.cell(row=start_row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=start_row, column=2).alignment = center

        headers = [first_col_header, "MT TTC", "M. H.T", "Taux TVA", "TVA"]
        for col_num, header in enumerate(headers, 2):
            cell = ws.cell(row=start_row + 1, column=col_num)
            cell.value = header
            cell.font = bold
            cell.alignment = center
            cell.border = border

        tva_total = 0
        for i, row in enumerate(data, start=start_row + 2):
            for j, key in enumerate(["Service", "TTC", "HT", "TVA Rate", "TVA"], 2):
                cell = ws.cell(row=i, column=j)
                if key == "TVA Rate":
                    cell.value = f"{row[key]}%"
                    cell.fill = grey_fill
                else:
                    cell.value = row[key]
                cell.border = border
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if key == "TVA":
                    tva_total += row[key]
                if row['Service'] == "Crédit Précédent":
                    cell.fill = red_fill
                elif row['Service'].upper().startswith("FAC"):
                    cell.fill = yellow_fill
                elif row['Service'].upper().startswith("FACTURE"):
                    cell.fill = red_fill

        total_row = start_row + 2 + len(data)
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=5)
        ws.cell(row=total_row, column=2).value = "la somme de TVA"
        ws.cell(row=total_row, column=2).fill = green_fill
        ws.cell(row=total_row, column=2).font = bold
        ws.cell(row=total_row, column=2).alignment = center
        ws.cell(row=total_row, column=6).value = round(tva_total, 2)
        ws.cell(row=total_row, column=6).font = bold
        ws.cell(row=total_row, column=6).fill = green_fill

        return tva_total, total_row

    clients = df[df['Role'] == 'Client'].to_dict(orient='records')
    fournisseurs = df[df['Role'] == 'Fournisseur'].to_dict(orient='records')

    credit_precedent_entries = [f for f in fournisseurs if f.get("Service") == "Crédit Précédent"]
    other_fournisseurs = [f for f in fournisseurs if f.get("Service") != "Crédit Précédent"]
    fournisseurs = other_fournisseurs + credit_precedent_entries

    start_row_clients = 3
    ca_title = f"C.A du {date_str}  {enterprise_name}"
    tva_client, end_row_clients = write_section(
        start_row_clients, ca_title, clients, is_client=True, first_col_header="Ventes")

    ws.cell(row=end_row_clients + 1, column=2).value = "Nombre de Facture"
    ws.cell(row=end_row_clients + 1, column=3).value = len(clients)
    ws.cell(row=end_row_clients + 1, column=2).font = bold
    ws.cell(row=end_row_clients + 1, column=2).fill = green_fill
    ws.cell(row=end_row_clients + 1, column=3).font = bold
    ws.cell(row=end_row_clients + 1, column=3).fill = green_fill

    start_row_fournisseurs = end_row_clients + 3
    tva_title = f"TVA RECUPERABLE le {date_str} "
    tva_fournisseur, end_row_fournisseurs = write_section(
        start_row_fournisseurs, tva_title, fournisseurs, is_client=False, first_col_header="Achats")

    final_row = end_row_fournisseurs + 3
    ws.merge_cells(start_row=final_row, start_column=2, end_row=final_row, end_column=5)
    ws.cell(row=final_row, column=2).value = "TVA DUE"
    ws.cell(row=final_row, column=2).font = Font(bold=True, color="FF0000")
    ws.cell(row=final_row, column=2).fill = red_fill
    ws.cell(row=final_row, column=2).alignment = center
    ws.cell(row=final_row, column=6).value = round(tva_client - tva_fournisseur, 2)
    ws.cell(row=final_row, column=6).font = Font(bold=True)

    towrite = io.BytesIO()
    wb.save(towrite)
    towrite.seek(0)
    return towrite

# 🔄 Now the main UI logic in a function
def show_tva_calculator():
    st.title("🧾 TVA Calculator & Excel Export")

    enterprise_name = st.text_input("Nom de l'entreprise")
    date_str = st.text_input("Date (MM/YYYY)", value="07/2025")

    if 'entries' not in st.session_state:
        st.session_state['entries'] = []

    with st.form("add_form"):
        col1, col2 = st.columns(2)
        with col1:
            role = st.selectbox("Type", ["Client", "Fournisseur", "Crédit Précédent"])
            service = st.text_input("Nom du service", value="")
        with col2:
            ttc = st.number_input("Montant TTC", min_value=0.0, step=0.01)
            tva_rate = st.number_input("Taux de TVA %", min_value=0.0, max_value=100.0, value=20.0)

        submitted = st.form_submit_button("Ajouter à la liste")
        if submitted and ttc > 0:
            if role == "Crédit Précédent":
                service = "Crédit Précédent"
                role_to_save = "Fournisseur"
            else:
                role_to_save = role
                if not service.strip():
                    next_id = len(st.session_state['entries']) + 1
                    service = f"Facture {next_id}" if role == "Fournisseur" else f"Service {next_id}"
            ht, tva = calculate_ht_tva(ttc, tva_rate)
            st.session_state['entries'].append({
                "Role": role_to_save,
                "Service": service,
                "TTC": ttc,
                "HT": ht,
                "TVA Rate": tva_rate,
                "TVA": tva
            })

    if st.session_state['entries']:
        df = pd.DataFrame(st.session_state['entries'])
        cols = st.columns(len(df.columns) + 1)
        cols[0].write("")
        for i, col_name in enumerate(df.columns):
            cols[i + 1].write(f"**{col_name}**")
        for idx, row in df.iterrows():
            cols = st.columns(len(df.columns) + 1)
            if cols[0].button("❌", key=f"del_{idx}"):
                st.session_state['entries'].pop(idx)
                st.rerun()
            for i, value in enumerate(row):
                cols[i + 1].write(value)

        if st.button("📤 Exporter vers Excel"):
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"tva_report_{enterprise_name or 'report'}_{current_time}.xlsx"
            towrite = export_to_excel(df, enterprise_name, date_str)
            st.download_button(
                label="📄 Télécharger Excel",
                data=towrite,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ Fichier Excel prêt à télécharger !")

    if st.button("🔁 Réinitialiser"):
        st.session_state['entries'] = []
